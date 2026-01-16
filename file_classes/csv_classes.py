
from openpyxl import load_workbook
from pathlib import Path

import re
from database_functions.schema import *


from sqlalchemy import *
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import declarative_base, sessionmaker, relationship



class rawScheduleEntry():
    def __init__(self, date:str, value:str, color:str = None):
        self.date = date
        self.value = value
        self.color = color

    def __repr__(self):
        return f"date: {self.date} | value: {self.value} | color: {self.color}"

class providerSchedule():
    def __init__(self, name:str):
        self.name = name
        self.schedule = []

    def __iter__(self):
        return iter(self.schedule)

    def add_entry(self, entry: rawScheduleEntry):
        if not isinstance(entry, rawScheduleEntry):
            raise ValueError("entries should be of type <rawScheduleEntry>")
        else:
            self.schedule.append(entry)


class ScheduleFile():
    def __init__(self, path, year_month=None, name_col = 1, date_row = 3, revsion_index = (1,2)):
        
        #Check for valid XLSX path
        path_obj = Path(path)
        if path_obj.exists() and path_obj.suffix.lower() == ".xlsx":
            self.path = path_obj
        else:
            raise FileNotFoundError

        self.data = self.read()

        self.name_col = name_col
        self.date_row = date_row
        self.revision_index = revsion_index

        self.bottom_bound = self.getBottomBound()
        self.right_bound = self.getRightBound()

        self.year_month = year_month if year_month else self.get_month_year_from_name()

        self.all_colors = {
            "rgb": set(),
            "indexed": set(),
            "theme": set()
        }

        self.all_cell_values = set()

    def get_month_year_from_name(self):
        
        MONTH_DICT = {"january" : "01",
                        "february" : "02",
                        "march" : "03",
                        "april" : "04",
                        "may" : " 05",
                        "june" : "06",
                        "july" : "07", 
                        "august" : "08", 
                        "september" : "09", 
                        "october" : "10", 
                        "november" : "11", 
                        "december" : "12"}
        
        for month in MONTH_DICT:
            if month in self.path.stem.lower():
                month_num_str = MONTH_DICT[month]
                break 
        
        YEAR_RE = re.compile(r"\b\d{4}\b")

        YEAR_RE = re.compile(r"(19|20)\d{2}")
        match = YEAR_RE.search(self.path.stem)
        year_num_str = match.group() if match else None

        if not year_num_str or not month_num_str:
            raise ValueError(f"Year: {year_num_str} or month {month_num_str} could not be extraxcted from {self.path.stem}")

        return f"{year_num_str}-{month_num_str}"

        

    def getRevisedDate(self):
        revision_cell_value = self.data.cell(column=self.revision_index[0], row=self.revision_index[1]).value
        cleaned_revision_cell_value = re.sub(r"[^0-9/]", "", revision_cell_value)
        day_month_year = cleaned_revision_cell_value.split('/')

        return f"{day_month_year[2]}-{day_month_year[0]}-{day_month_year[1]}"
    
    # def getYearMonth(self):
    #     return self.getRevisedDate()[:7]

    def read(self):
        wb = load_workbook(self.path, data_only=True)
        self.data = wb.active
        return wb.active
    
    def getBottomBound(self):
        
        current_max = self.date_row+1

        #Yields cells from name colum starting under the date row
        name_column = self.data.iter_cols(
            min_col=self.name_col, 
            max_col=self.name_col, 
            min_row=self.date_row+1) 
        
        for cell_tuple in name_column:
            for cell in cell_tuple:
                if "#" in cell.value:
                    self.bottom_bound = current_max
                    return current_max - 1
                else:
                    current_max += 1

        return self.data.max_row
    
    def getRightBound(self):
        current_max = self.name_col

        date_row = self.data.iter_rows(
            min_col=self.name_col+1, 
            min_row=self.date_row,
            max_row=self.date_row)
        
        for cell_tuple in date_row:
            for cell in cell_tuple:
                if str(cell.value).isdigit():
                    current_max += 1
                else:
                    return current_max
        return self.data.max_column

    
    def getProviderRows(self):

        providers = dict()

        name_column = self.data.iter_cols(
            min_col=self.name_col, 
            max_col=self.name_col, 
            min_row=self.date_row+1,
            max_row=self.bottom_bound)
        
        for cell_tuple in name_column:
            for cell in cell_tuple:
                providers[cell.value] = cell.row

        return providers
    
    
    def getProviderSchedules(self):

        schedules = dict()
        provider_rows = self.getProviderRows() 
        for provider in provider_rows:

            row_num = provider_rows[provider]
            provider_schedule = providerSchedule(provider)

            provider_row = self.data.iter_rows(
                min_col=self.name_col+1,
                max_col=self.right_bound,
                min_row=row_num,
                max_row=row_num)

            for cell_tuple in provider_row:
                for cell in cell_tuple:
                    cell_date = self.year_month+"-"+f"{self.getAssociatedDate(cell.column)['date']:02}"
                    cell_data = rawScheduleEntry(cell_date, cell.value, self.getCellColor(cell))
                    provider_schedule.add_entry(cell_data)
            
            schedules[provider] = provider_schedule
        return schedules
    
    def getAllCellValues(self):
        all_values = set()
        provider_rows = self.getProviderRows()
        for provider in provider_rows:

            row_num = provider_rows[provider]

            provider_row = self.data.iter_rows(
                min_col=self.name_col+1,
                max_col=self.right_bound,
                min_row=row_num,
                max_row=row_num
                )

            for cell_tuple in provider_row:
                for cell in cell_tuple:
                    all_values.add(str(cell.value))
        
        return all_values


    def getAssociatedDate(self, col:int) -> dict:
        '''
        Docstring for getAssociatedDate
        :param col: col of the cell to get associated date from
        :type col: int
        :return: returns the date and day associated with a cell
        :rtype: dict
        '''

        ALLOWED_DAYS = set(["MON","TUE","WED","THU","FRI"])

        date = self.data.cell(row=self.date_row, column=col).value
        day = self.data.cell(row=self.date_row-1, column=col).value
        if day not in ALLOWED_DAYS:
            raise ValueError(f"{col, self.date_row} Attempting to access day {day} is not in {ALLOWED_DAYS}")
        return {"date":date, "day":day}
    
    def getCellColor(self, cell):
        fill = cell.fill

        if not fill or fill.fill_type != "solid":
            return None
        
        color = fill.start_color

        if color is None:
            return None
        
        if color.type == "rgb" and color.rgb:
            self.all_colors[color.type].add(color.rgb[-6:])
            return "#" + color.rgb[-6:] #remove alpha values
        
        else:
            return "#808080"
        

from datetime import datetime
from dateutil.relativedelta import relativedelta

def previous_year_month(iso_year_month: str) -> str:
    dt = datetime.strptime(iso_year_month, "%Y-%m")
    return (dt - relativedelta(months=1)).strftime("%Y-%m")




# def get_providers(session, 
#     date: str = None, 
#     names_excluded: list = None,
#     colors_excluded: list = None,
#     values_excluded: list = None,
#     names_specified: list = None,
#     colors_specified: list = None,
#     values_specified: list = None,

#     getNames: bool = False
#     ):
        
#     filters = []

#     if date:
#         filters.append(ProviderDate.date == date)

#     #Name Filtering
#     if names_excluded and names_specified:
#         raise NameError("Cannot exclude and specify names")
#     elif names_excluded:
#         filters.append(Provider.name.notin_(names_excluded))
#     elif names_specified:
#         filters.append(Provider.name.in_(names_specified))

#     #Color Filtering
#     if colors_excluded and colors_specified:
#         raise NameError("Cannot exclude and specify colors")
#     elif colors_excluded:
#         filters.append(ProviderDate.color.notin_(colors_excluded))
#     elif colors_specified:
#         filters.append(ProviderDate.color.in_(colors_specified))

#     #Value Filtering
#     if values_excluded and values_specified:
#         raise NameError("Cannot exclude and specify values")
#     elif values_excluded:
#         filters.append(ProviderDate.value.notin_(values_excluded))
#     elif values_specified:
#         filters.append(ProviderDate.value.in_(values_specified))

#     query = (select(Provider)
#              .join(ProviderDate)
#              .where(and_(*filters)))
    
#     results = session.execute(query).scalars().all()
    
#     if getNames:
#         return [p.name for p in results]
#     else:
#         return results

    



class RankedColumn():
    def __init__(self, lists: list, spacing: int = 1):
        self.lists = lists
        self.spacing = spacing #Vertical

class RankedDay():
    def __init__(self, list_columns: RankedColumn, spacing: int = 1):
        self.list_columns = list_columns
        self.spacing = spacing #Horizontal

class RankedRow():
    def __init__(self):
        pass

